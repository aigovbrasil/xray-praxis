#!/usr/bin/env python3
"""
PRAXIS — generate_executive_xls.py

Generate A-OPS executive spreadsheet using openpyxl with PRAXIS
design tokens applied (GAP-08 resolved).

Usage:
    python generate_executive_xls.py [--manifest <path>] [--output <path>]
                                     [--data <path>]
"""

import argparse
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    import yaml
except ImportError:
    print("ERROR: pyyaml not installed.", file=sys.stderr)
    sys.exit(2)

try:
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed.", file=sys.stderr)
    sys.exit(2)


# Design tokens (GAP-08)
COLOR_PRIMARY = "1B2A4A"
COLOR_SURFACE = "F5F7FA"
COLOR_DIVIDER = "D0D7E0"
COLOR_WHITE = "FFFFFF"

FONT_FAMILY = "Inter"
FONT_HEADER_PT = 10
FONT_BODY_PT = 10
ROW_HEIGHT = 18
COL_MIN_WIDTH = 15

HEADER_FILL = PatternFill("solid", fgColor=COLOR_PRIMARY)
HEADER_FONT = Font(name=FONT_FAMILY, size=FONT_HEADER_PT, bold=True,
                   color=COLOR_WHITE)
ALT_FILL = PatternFill("solid", fgColor=COLOR_SURFACE)
WHITE_FILL = PatternFill("solid", fgColor=COLOR_WHITE)
THIN_DIVIDER = Side(border_style="thin", color=COLOR_DIVIDER)
BORDER_ALL = Border(left=THIN_DIVIDER, right=THIN_DIVIDER,
                    top=THIN_DIVIDER, bottom=THIN_DIVIDER)
BODY_FONT = Font(name=FONT_FAMILY, size=FONT_BODY_PT)


def set_columns(ws, widths: list[int]):
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(w, COL_MIN_WIDTH)


def write_header_row(ws, row_idx: int, headers: list[str]):
    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                   wrap_text=True)
        cell.border = BORDER_ALL
    ws.row_dimensions[row_idx].height = ROW_HEIGHT


def write_data_row(ws, row_idx: int, values: list, is_alt: bool):
    fill = ALT_FILL if is_alt else WHITE_FILL
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = BODY_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                   wrap_text=True)
        cell.border = BORDER_ALL
    ws.row_dimensions[row_idx].height = ROW_HEIGHT


def build_resumo(ws, manifest: dict):
    ws.title = "Resumo Executivo"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    title_cell = ws.cell(row=1, column=1, value="Resumo Executivo")
    title_cell.font = Font(name=FONT_FAMILY, size=14, bold=True,
                           color=COLOR_PRIMARY)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    headers = ["Campo", "Valor"]
    write_header_row(ws, 3, headers)
    rows = [
        ("case_id", manifest.get("case_id", "")),
        ("consultant_id", manifest.get("consultant_id", "")),
        ("client", manifest.get("client_identity", {}).get("company_name", "")),
        ("scenario", manifest.get("scenario", "")),
        ("current_phase", manifest.get("current_phase", "")),
        ("created_at", manifest.get("created_at", "")),
        ("updated_at", manifest.get("updated_at", "")),
        ("artifacts_count", len(manifest.get("artifacts_produced", []))),
        ("gates_passed_count", len(manifest.get("gates_passed", []))),
    ]
    for i, (k, v) in enumerate(rows):
        write_data_row(ws, 4 + i, [k, v], is_alt=(i % 2 == 1))
    set_columns(ws, [25, 50])


def build_diagnostico(wb, payload: dict):
    ws = wb.create_sheet("Diagnóstico")
    headers = ["Problema", "Causa-raiz", "Rótulo", "Score prioridade"]
    write_header_row(ws, 1, headers)
    rows = payload.get("diagnostic_rows", [])
    if not rows:
        rows = [
            ("[A PREENCHER]", "[A PREENCHER]", "[HIPÓTESE]", "")
        ]
    for i, row in enumerate(rows):
        write_data_row(ws, 2 + i, list(row), is_alt=(i % 2 == 1))
    set_columns(ws, [40, 40, 18, 18])


def build_plano_acao(wb, payload: dict):
    ws = wb.create_sheet("Plano de Ação")
    headers = ["What", "Why", "Where", "When", "Who", "How", "How much"]
    write_header_row(ws, 1, headers)
    rows = payload.get("action_rows", [])
    if not rows:
        rows = [("[A PREENCHER]",) * 7]
    for i, row in enumerate(rows):
        write_data_row(ws, 2 + i, list(row), is_alt=(i % 2 == 1))
    set_columns(ws, [25, 25, 18, 18, 20, 25, 18])


def build_simulacao(wb, payload: dict):
    ws = wb.create_sheet("Simulação")
    headers = ["Métrica", "Conservador", "Base", "Otimista", "Rótulo"]
    write_header_row(ws, 1, headers)
    rows = payload.get("scenario_rows", [])
    if not rows:
        rows = [("[A PREENCHER]", "", "", "", "[HIPÓTESE]")]
    for i, row in enumerate(rows):
        write_data_row(ws, 2 + i, list(row), is_alt=(i % 2 == 1))
    set_columns(ws, [30, 18, 18, 18, 18])


def build_proximos_passos(wb, payload: dict):
    ws = wb.create_sheet("Próximos Passos")
    headers = ["Ação", "Owner", "Prazo", "KPI", "Status"]
    write_header_row(ws, 1, headers)
    rows = payload.get("next_steps_rows", [])
    if not rows:
        rows = [("[A PREENCHER]", "[A DEFINIR]", "[A DEFINIR]", "[A DEFINIR]",
                "pendente")]
    for i, row in enumerate(rows):
        write_data_row(ws, 2 + i, list(row), is_alt=(i % 2 == 1))
    set_columns(ws, [40, 25, 18, 25, 15])


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate A-OPS XLSX.")
    parser.add_argument("--manifest", default="manifest.yaml")
    parser.add_argument("--output", default="A-OPS.xlsx")
    parser.add_argument("--data", default=None,
                        help="Optional YAML/JSON file with diagnostic_rows, "
                             "action_rows, scenario_rows, next_steps_rows")
    args = parser.parse_args()

    manifest_path = Path(args.manifest)
    if not manifest_path.exists():
        print(f"ERROR: manifest not found at {manifest_path}", file=sys.stderr)
        return 2

    with manifest_path.open("r", encoding="utf-8") as f:
        manifest = yaml.safe_load(f)

    payload: dict = {}
    if args.data:
        data_path = Path(args.data)
        if data_path.exists():
            with data_path.open("r", encoding="utf-8") as f:
                payload = yaml.safe_load(f) or {}

    wb = Workbook()
    build_resumo(wb.active, manifest)
    build_diagnostico(wb, payload)
    build_plano_acao(wb, payload)
    build_simulacao(wb, payload)
    build_proximos_passos(wb, payload)

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"A-OPS written to: {out_path}")
    print(f"Generated at: {datetime.now(timezone.utc).isoformat()}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
